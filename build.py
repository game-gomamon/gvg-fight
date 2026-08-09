#!/usr/bin/env python3
"""
Rebuild the Guild Fight site data from etheria_restart.xlsx.

Usage:
    python3 build.py                      # looks for etheria_restart.xlsx next to this file
    python3 build.py path/to/book.xlsx

Writes:
    data.json              lineups + roster (this is what the page loads)
    assets/animus/*.webp   portraits pulled out of the Information sheet
    assets/shell/*.webp

Requires: pip install openpyxl pillow lxml
"""

import json
import os
import re
import shutil
import sys
import zipfile
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook
from PIL import Image

HERE = Path(__file__).resolve().parent
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else HERE / "etheria_restart.xlsx"

INFO_SHEET = "Information"
FIGHT_SHEET = "Guild_fight"
ANIMUS_NAME_COL, ANIMUS_IMG_COL = 1, 2   # A, B
SHELL_NAME_COL, SHELL_IMG_COL = 4, 5     # D, E
PORTRAIT_PX = 160
WEBP_QUALITY = 82

REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"


def slug(name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return s or "unknown"


def extract_cell_images(xlsx: Path, tmp: Path):
    """Map 'Sheet1!B2' style refs to the image bytes of Excel's in-cell pictures.

    Excel stores in-cell images as "rich values": the cell carries a vm= index,
    which points into rdrichvalue.xml, which points into richValueRel.xml,
    which finally resolves to a file in xl/media/.
    """
    with zipfile.ZipFile(xlsx) as z:
        z.extractall(tmp)

    rv_path = tmp / "xl/richData/rdrichvalue.xml"
    if not rv_path.exists():
        return {}, tmp

    rv = etree.parse(str(rv_path)).getroot()
    rv_to_rel = [int(el.find("{*}v").text) for el in rv]

    rvr = etree.parse(str(tmp / "xl/richData/richValueRel.xml")).getroot()
    rel_ids = [el.get(REL_NS) for el in rvr]

    rels = etree.parse(str(tmp / "xl/richData/_rels/richValueRel.xml.rels")).getroot()
    rid_to_file = {el.get("Id"): os.path.basename(el.get("Target")) for el in rels}

    # cell ref -> media filename, for the Information sheet only
    sheet = etree.parse(str(tmp / "xl/worksheets/sheet1.xml"))
    cell_to_media = {}
    for c in sheet.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c"):
        vm = c.get("vm")
        if not vm:
            continue
        try:
            cell_to_media[c.get("r")] = rid_to_file[rel_ids[rv_to_rel[int(vm) - 1]]]
        except (IndexError, KeyError):
            pass
    return cell_to_media, tmp


def save_portrait(src: Path, dest: Path):
    img = Image.open(src).convert("RGBA")
    if max(img.size) != PORTRAIT_PX:
        img = img.resize((PORTRAIT_PX, PORTRAIT_PX), Image.LANCZOS)
    dest.parent.mkdir(parents=True, exist_ok=True)
    img.save(dest, "WEBP", quality=WEBP_QUALITY, method=6)


def main():
    if not XLSX.exists():
        sys.exit(f"Can't find {XLSX}. Put the workbook next to build.py, or pass its path.")

    tmp = HERE / ".xlsx_unpacked"
    if tmp.exists():
        shutil.rmtree(tmp)
    cell_to_media, tmp = extract_cell_images(XLSX, tmp)
    media_dir = tmp / "xl/media"

    wb = load_workbook(XLSX, data_only=True)
    info = wb[INFO_SHEET]

    roster = {"animus": {}, "shell": {}}
    for row in range(2, info.max_row + 1):
        for kind, name_col, img_col, folder in (
            ("animus", ANIMUS_NAME_COL, ANIMUS_IMG_COL, "animus"),
            ("shell", SHELL_NAME_COL, SHELL_IMG_COL, "shell"),
        ):
            name = info.cell(row, name_col).value
            if not name:
                continue
            name = str(name).strip()
            key = slug(name)
            entry = {"name": name, "img": None}
            ref = f"{info.cell(row, img_col).coordinate}"
            media = cell_to_media.get(ref)
            if media and (media_dir / media).exists():
                out = HERE / "assets" / folder / f"{key}.webp"
                save_portrait(media_dir / media, out)
                entry["img"] = f"assets/{folder}/{key}.webp"
            roster[kind][key] = entry

    fight = wb[FIGHT_SHEET]
    lineups = []
    for row in fight.iter_rows(min_row=2, values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in list(row) + [""] * 13]
        if not cells[0]:
            continue
        enemy = [{"animus": cells[i], "shell": cells[i + 1]} for i in (0, 2, 4)]
        ours = [{"animus": cells[i], "shell": cells[i + 1]} for i in (6, 8, 10)]
        lineups.append({"enemy": enemy, "ours": ours, "guide": cells[12]})

    data = {
        "source": XLSX.name,
        "roster": roster,
        "lineups": lineups,
    }
    (HERE / "data.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8"
    )

    shutil.rmtree(tmp, ignore_errors=True)
    print(
        f"Wrote data.json - {len(lineups)} lineups, "
        f"{len(roster['animus'])} animus, {len(roster['shell'])} shell portraits."
    )


if __name__ == "__main__":
    main()
